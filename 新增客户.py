import time

import openpyxl
import pyautogui
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


def input_txt(path, text):
    input_box = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located(
            (By.XPATH, path))
    )
    input_box.clear()
    input_box.send_keys(text)


def button_click(path):
    button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.XPATH, path))
    )
    button.click()
    time.sleep(2)


def select(path, value):
    value_int = int(value)
    select = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, path))
    )
    select.click()
    time.sleep(0.5)
    # 创建 ActionChains 对象
    # 模拟按下下箭头键
    while value_int > 0:
        pyautogui.press('down')
        value_int = value_int - 1
        time.sleep(0.5)

    # 模拟按下回车键
    pyautogui.press('enter')


driver = webdriver.Chrome()
driver.get('http://192.168.20.61')
#
# # 登录
input_txt("/html/body/div[1]/div[1]/div/div/div[2]/form[1]/div/div[3]/div/div/div/div/input", "A0004")
button_click("/html/body/div[1]/div[1]/div/div/div[2]/form[1]/div/div[6]/div/div/button")
#
# # 进入我的客户列表
driver.get('http://192.168.20.61/custom/myCustom')

# 打开 Excel 文件
wb = openpyxl.load_workbook('D:\desk\客户信息.xlsx')

# 客户信息
sheet1 = wb.worksheets[0]
sheet1_max_row = sheet1.max_row
sheet1_max_column = sheet1.max_column
# 联系人信息
sheet2 = wb.worksheets[1]
sheet2_max_row = sheet2.max_row
sheet2_max_column = sheet2.max_column

# 从第5行开始遍历每一列
start_row = 5
start_col = 2

# 判断是否需要操作该信息
is_oper_row = 1
# 判断使用哪个方法
use_method_row = 2
# XPath
use_XPath_row = 3

for row in range(start_row, sheet1_max_row + 1):
    # 点击新增
    button_click("/html/body/div[1]/section/div[2]/div/div[1]/div/section/div[1]/div/div/form/div[3]/div/button[3]")
    for col in range(start_col, sheet1_max_column + 1):
        # 先判断当前列是否需要操作
        is_oper = sheet1.cell(row=is_oper_row, column=col).value
        if is_oper == "1":
            value = sheet1.cell(row=row, column=col).value
            # 获取使用的方法
            use_method = sheet1.cell(row=use_method_row, column=col).value
            use_XPath = sheet1.cell(row=use_XPath_row, column=col).value
            if use_method == "input":
                input_txt(use_XPath, value)
            elif use_method == "select":
                select(use_XPath, value)

    # 点击新增联系人
    button_click("/html/body/div[1]/section/div[2]/div/div[1]/div/section/div/div/div/form/div[2]/div[2]/div/button")

    for col1 in range(start_col, sheet2_max_column + 1):
        # 先判断当前列是否需要操作
        is_oper = sheet2.cell(row=is_oper_row, column=col1).value

        if is_oper == "1":
            value = sheet2.cell(row=row, column=col1).value

            # 获取使用的方法
            use_method = sheet2.cell(row=use_method_row, column=col1).value
            use_XPath = sheet2.cell(row=use_XPath_row, column=col1).value
            if use_method == "input":
                input_txt(use_XPath, value)
            elif use_method == "select":
                select(use_XPath, value)
    button_click("/html/body/div[1]/section/div[2]/div/div[1]/div/section/div/div/div/div[2]/div/div/footer/button[2]")
    button_click("/html/body/div[1]/section/div[2]/div/div[1]/div/section/div/div/div/div[1]/button[2]")

# 关闭 Excel 文件
wb.close()
